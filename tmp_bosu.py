from builtins import print
from bs4 import BeautifulSoup

from utils import get_driver, close_driver, get_unit, get_table

import sys
import time
import re

from openpyxl import Workbook, load_workbook


excel_file_name = "2016_보수한도.xlsx"
excel_sheet_title = "2016 보수한도"

# dart 문서번호
def get_rcpNo(jm_code, keywod, st_dt, en_dt):
    driver = get_driver('C:\\Users\\admin\\PycharmProjects\\webCrawl\\chromedriver.exe',
                        'http://dart.fss.or.kr/dsab002/main.do#')                                   # 드라이버 로드

    driver.find_element_by_name('textCrpNm').send_keys(jm_code)                                     # 종목코드
    driver.find_element_by_xpath('//*[@id="startDate"]').send_keys(st_dt)                           # 기간 시작
    driver.find_element_by_xpath('//*[@id="endDate"]').send_keys(en_dt)                             # 기간 종료
    driver.find_element_by_id('reportName').send_keys(keywod)                                       # 검색구분 : 결의

    driver.find_element_by_xpath('//*[@id="searchForm"]/fieldset/div/p[8]/input').click()           # 검색

    res_list = driver.find_elements_by_xpath('//*[@id="listContents"]/div[1]/table/tbody/tr')       # 결과 리스트

    # 최상위 데이터만 수집 ( => 짧은 주기로 수집해야 함)
    # 결과 리스트에서 가용 데이터 추출
    if len(res_list) == 0:
        print('검색 결과가 없습니다.')
        sys.exit(0)
    else:
        item = res_list[0]
    # 문서번호
    rcp_no = item.find_elements_by_tag_name('td')[2].find_element_by_tag_name('a').get_attribute('href')[-14:]
    # 기재정정
    rcp_yn = item.find_elements_by_tag_name('td')[2].find_element_by_tag_name('a').find_element_by_tag_name('span').text

    if len(rcp_no) != 14:
        print('rcpNo 형식이 다릅니다.')
        sys.exit(0)

    if '첨부' in rcp_yn:
        print('첨부정정은 수집대상 제외')
        sys.exit(0)

    close_driver(driver)

    return rcp_no, rcp_yn


# 주총소집공고 내용
def get_hando(driver, jm_code, gijun_yy):
    if ('임원 및 직원' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[10]/div/a').text and
            '관한 사항' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[10]/div/a').text):
        driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[10]/ul/li[2]/div/a').click()

    elif ('임원 및 직원' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[11]/div/a').text and
            '관한 사항' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[11]/div/a').text):
        driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[11]/ul/li[2]/div/a').click()

    elif ('임원 및 직원' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[12]/div/a').text and
            '관한 사항' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[12]/div/a').text):
        driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[12]/ul/li[2]/div/a').click()

    elif ('임원 및 직원' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[13]/div/a').text and
          '관한 사항' in driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[13]/div/a').text):
        driver.find_element_by_xpath('//*[@id="ext-gen10"]/div/li[13]/ul/li[2]/div/a').click()

    left_navi = driver.find_elements_by_xpath('//*[@id="ext-gen10"]/div/li')

    driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))           # iframe 가져오기

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    ps = soup.find_all('p')
    tables = soup.find_all('table')

    hando_gb = re.compile("구[ \s]*분")
    hando_cnt = re.compile("인[ \s]*원[ \s]*수|인[ \s]*원")
    hando_amt = re.compile("주주총회[ \s]*승인[ \s]*금액|주주[ \s]*총회[ \s]*승인[ \s]*금액")

    jigup_gb = re.compile("구[ \s]*분")
    jigup_cnt = re.compile("인[ \s]*원[ \s]*수|인[ \s]*원")
    jigup_tot = re.compile("보수[ \s]*총액|보[ \s]*수[ \s]*총[ \s]*액")
    jigup_per = re.compile("1인당[ \s]*평균보수액|1[ \s]*인당[ \s]*평균보수액|1[ \s]*인당[ \s]*평균[ \s]*보수액")

    pttn_del = re.compile("합[ \S]*계|[ \S]*계")
    pttn_num = re.compile("[0-9]")

    cnt = 0
    hando_num = 0
    for table in tables:
        if hando_gb.search(table.text) and hando_cnt.search(table.text) and hando_amt.search(table.text):
            hando_num = cnt
        cnt = cnt + 1

    hando_arr = []
    hando_unit = 1
    if hando_num > 0:
        tmp_tb = get_table(tables[hando_num])
        d_tmp_tb = []
        for d in range(0, len(tmp_tb)):
            if (not pttn_del.search(tmp_tb[d][0]) and not pttn_del.search(tmp_tb[d][1])
                    and not pttn_del.search(tmp_tb[d][2]) and not pttn_del.search(tmp_tb[d][3])):
                d_tmp_tb.append(tmp_tb[d])

        hando_arr.extend(d_tmp_tb)
        hando_unit = get_unit(tables[hando_num - 1])

    for i in range(0, len(hando_arr)):
        tmp_amt = "".join(pttn_num.findall(hando_arr[i][2]))
        if not tmp_amt:
            tmp_amt = "0"

        if "백만" in hando_unit:
            hando_arr[i][2] = int(tmp_amt) * 1000
        elif "천" in hando_unit:
            hando_arr[i][2] = int(tmp_amt)
        elif "억" in hando_unit:
            hando_arr[i][2] = int(tmp_amt) * 100000
        elif hando_unit == "원":
            hando_arr[i][2] = round(int(tmp_amt) / 1000)

        hando_arr[i].insert(0, jm_code)
        hando_arr[i].insert(0, gijun_yy)

    return hando_arr


# 주총공고수집
def hando_main(jm_code, rcp_no, gijun_yy):
    # driver 세팅
    driver = get_driver('C:\\Users\\admin\\PycharmProjects\\webCrawl\\chromedriver.exe',
                        'http://dart.fss.or.kr/dsaf001/main.do?rcpNo={0}'.format(rcp_no))

    hando = []
    hando.extend(get_hando(driver, jm_code, gijun_yy))

    for h in hando:
        print(h)

    # driver close
    close_driver(driver)


# html &nbsp; 삭제
def del_entity(ent):
    ent = ent.replace(u'\xa0', u' ').strip()

    return ent


def make_excel():
    work_book = Workbook()
    sheet1 = work_book.active
    sheet1.title = excel_sheet_title

    sheet1.cell(row=1, column=1).value = "회계년도"
    sheet1.cell(row=1, column=2).value = "종목코드"
    sheet1.cell(row=1, column=3).value = "이사구분"
    sheet1.cell(row=1, column=4).value = "인원수"
    sheet1.cell(row=1, column=5).value = "한도금액"
    sheet1.cell(row=1, column=6).value = "비고"

    work_book.save(filename=excel_file_name)
    work_book.close()


def insert_data_excel(crawling_results):
    excel_file = load_workbook(excel_file_name)
    sheet1 = excel_file[excel_sheet_title]

    excel_row = len(crawling_results)
    excel_column = 6
    for data in crawling_results:
        sheet1.cell(row=excel_row, column=excel_column).value = data


if __name__== "__main__":
    yy = time.strftime("%Y")
    rcp_no_gong, rcp_yn_gong = get_rcpNo('006200', '사업보고서', '20160101', '20161231')

    hando_main('006200', rcp_no_gong, '2016')

